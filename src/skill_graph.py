"""
ESCO–O*NET Hybrid Knowledge Graph Module
=========================================
In-memory graph for skill hierarchy, occupation-skill relationships,
and technology tools. Used for graph-based query expansion in the
search pipeline.

Data sources:
  - data/esco/skill_broader_relations.csv   (ESCO skill hierarchy)
  - data/esco/occupation_skills.csv          (ESCO occupation → skills)
  - data/onet/ESCO_to_ONET-SOC.xlsx         (ESCO ↔ O*NET crosswalk)
  - data/onet/Technology_Skills.xlsx         (O*NET technology tools)

Usage:
  from skill_graph import get_skill_graph
  graph = get_skill_graph()
  expanded = graph.expand_query_terms(["Python", "SQL"], max_terms=15)
"""

import csv
import os
import logging
import time
from collections import defaultdict

logger = logging.getLogger(__name__)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "..", "data")

_graph_instance = None


class SkillGraph:
    """In-memory Hybrid Knowledge Graph for ESCO skills, occupations, and O*NET tools."""

    def __init__(self, data_dir=None):
        self.data_dir = data_dir or DATA_DIR

        # Node storage: uri -> {title, type}
        #   type: "SKILL" | "OCCUPATION" | "TOOL"
        self.nodes = {}

        # ESCO adjacency lists
        self.broader = defaultdict(set)       # skill -> broader skill URIs
        self.narrower = defaultdict(set)      # skill -> narrower skill URIs
        self.occ_essential = defaultdict(set) # occupation -> essential skill URIs
        self.occ_optional = defaultdict(set)  # occupation -> optional skill URIs
        self.skill_to_occs = defaultdict(set) # skill -> occupation URIs

        # O*NET adjacency lists
        self.occ_tools = defaultdict(set)     # occupation_uri -> tool URIs
        self.tool_to_occs = defaultdict(set)  # tool_uri -> occupation URIs

        # ESCO ↔ O*NET crosswalk: esco_occ_uri -> onet_code
        self.esco_to_onet = {}
        self.onet_to_esco = defaultdict(set)

        # Title-based index (lowercase → URI)
        self.title_to_uri = {}

        # Stats
        self._loaded = False

    def load(self):
        """Load graph data from ESCO CSVs and O*NET XLSX files."""
        if self._loaded:
            return

        t0 = time.perf_counter()

        hierarchy_file = os.path.join(self.data_dir, "esco", "skill_broader_relations.csv")
        occ_skills_file = os.path.join(self.data_dir, "esco", "occupation_skills.csv")
        crosswalk_file = os.path.join(self.data_dir, "onet", "ESCO_to_ONET-SOC.xlsx")
        tech_skills_file = os.path.join(self.data_dir, "onet", "Technology_Skills.xlsx")

        if not os.path.exists(hierarchy_file):
            logger.warning(f"Skill hierarchy not found: {hierarchy_file}")
            return

        # ── Layer 1: ESCO Skill Hierarchy ──
        h_count = 0
        with open(hierarchy_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                skill_uri = row["skill_uri"]
                skill_title = row["skill_title"]
                broader_uri = row.get("broader_uri", "")
                broader_title = row.get("broader_title", "")

                # Add skill node
                if skill_uri not in self.nodes:
                    self.nodes[skill_uri] = {"title": skill_title, "type": "SKILL"}
                    self.title_to_uri[skill_title.lower()] = skill_uri

                # Add broader relation
                if broader_uri and broader_title != "(top-level)":
                    if broader_uri not in self.nodes:
                        self.nodes[broader_uri] = {"title": broader_title, "type": "SKILL"}
                        self.title_to_uri[broader_title.lower()] = broader_uri

                    self.broader[skill_uri].add(broader_uri)
                    self.narrower[broader_uri].add(skill_uri)
                    h_count += 1

        # ── Layer 2: ESCO Occupation-Skill (bipartite) ──
        o_count = 0
        if os.path.exists(occ_skills_file):
            with open(occ_skills_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    occ_uri = row["occupation_uri"]
                    occ_title = row["occupation_title"]
                    skill_uri = row["skill_uri"]
                    skill_title = row["skill_title"]
                    rel_type = row["relation_type"]

                    if occ_uri not in self.nodes:
                        self.nodes[occ_uri] = {"title": occ_title, "type": "OCCUPATION"}
                        self.title_to_uri[occ_title.lower()] = occ_uri

                    if skill_uri not in self.nodes:
                        self.nodes[skill_uri] = {"title": skill_title, "type": "SKILL"}
                        self.title_to_uri[skill_title.lower()] = skill_uri

                    if rel_type == "essential":
                        self.occ_essential[occ_uri].add(skill_uri)
                    else:
                        self.occ_optional[occ_uri].add(skill_uri)
                    self.skill_to_occs[skill_uri].add(occ_uri)
                    o_count += 1

        # ── Layer 3: O*NET Technology Tools (via crosswalk) ──
        t_count = self._load_onet_tools(crosswalk_file, tech_skills_file)

        elapsed = time.perf_counter() - t0
        skills = sum(1 for n in self.nodes.values() if n["type"] == "SKILL")
        occs = sum(1 for n in self.nodes.values() if n["type"] == "OCCUPATION")
        tools = sum(1 for n in self.nodes.values() if n["type"] == "TOOL")

        logger.info(
            f"SkillGraph loaded: {skills} skills, {occs} occupations, {tools} tools, "
            f"{h_count} hierarchy + {o_count} occ-skill + {t_count} occ-tool edges "
            f"in {elapsed:.2f}s"
        )
        self._loaded = True

    def _load_onet_tools(self, crosswalk_file, tech_skills_file):
        """Load O*NET Technology Tools and link to ESCO occupations via crosswalk."""
        if not os.path.exists(crosswalk_file) or not os.path.exists(tech_skills_file):
            logger.info("O*NET files not found, skipping technology tools layer")
            return 0

        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed, skipping O*NET integration")
            return 0

        # Step 1: Build crosswalk — ESCO occupation title → O*NET SOC code
        occ_title_to_onet = {}  # esco_occ_title(lower) -> onet_code
        wb = openpyxl.load_workbook(crosswalk_file, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=5, values_only=True):  # Skip header rows
            if row[0] and row[2]:
                esco_title = str(row[1]).strip().lower() if row[1] else ""
                onet_code = str(row[2]).strip()
                if esco_title:
                    occ_title_to_onet[esco_title] = onet_code
        wb.close()

        # Map ESCO occupation URIs to O*NET codes
        for uri, node in self.nodes.items():
            if node["type"] == "OCCUPATION":
                onet_code = occ_title_to_onet.get(node["title"].lower())
                if onet_code:
                    self.esco_to_onet[uri] = onet_code
                    self.onet_to_esco[onet_code].add(uri)

        # Step 2: Load technology tools and link to mapped occupations
        #   Build O*NET code → set of tool names
        onet_code_tools = defaultdict(set)  # onet_code -> {tool_name, ...}
        wb = openpyxl.load_workbook(tech_skills_file, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            onet_code = str(row[0]).strip() if row[0] else ""
            tool_name = str(row[2]).strip() if row[2] else ""
            if onet_code and tool_name:
                onet_code_tools[onet_code].add(tool_name)
        wb.close()

        # Step 3: Create TOOL nodes and occ→tool edges
        t_count = 0
        for esco_uri, onet_code in self.esco_to_onet.items():
            for tool_name in onet_code_tools.get(onet_code, set()):
                tool_uri = f"onet:tool:{tool_name.lower().replace(' ', '_')}"

                if tool_uri not in self.nodes:
                    self.nodes[tool_uri] = {"title": tool_name, "type": "TOOL"}
                    # Only index in title_to_uri if no SKILL/OCCUPATION already uses this name
                    key = tool_name.lower()
                    existing_uri = self.title_to_uri.get(key)
                    if not existing_uri or self.nodes.get(existing_uri, {}).get("type") == "TOOL":
                        self.title_to_uri[key] = tool_uri

                self.occ_tools[esco_uri].add(tool_uri)
                self.tool_to_occs[tool_uri].add(esco_uri)
                t_count += 1

        crosswalk_count = len(self.esco_to_onet)
        logger.info(
            f"  O*NET layer: {crosswalk_count} occupations crosswalked, "
            f"{sum(1 for n in self.nodes.values() if n['type'] == 'TOOL')} tools loaded"
        )
        return t_count

    def find_uri(self, title, prefer_skill=True):
        """Find URI by title (case-insensitive, exact match).
        When prefer_skill=True, prioritizes SKILL/OCCUPATION over TOOL nodes."""
        key = title.lower().strip()

        # Try with common ESCO suffixes first (to prioritize SKILL over TOOL)
        for suffix in ["(computer programming)", "(software)"]:
            variant = f"{key} {suffix}"
            if variant in self.title_to_uri:
                return self.title_to_uri[variant]

        # Direct lookup
        if key in self.title_to_uri:
            uri = self.title_to_uri[key]
            # If we prefer skills and found a TOOL, check if there's also a SKILL
            if prefer_skill and self.nodes.get(uri, {}).get("type") == "TOOL":
                # Search for a SKILL node with same base name
                for suffix in ["(computer programming)", "(software)", ""]:
                    variant = f"{key} {suffix}".strip() if suffix else key
                    candidate = self.title_to_uri.get(variant)
                    if candidate and self.nodes.get(candidate, {}).get("type") == "SKILL":
                        return candidate
            return uri

        return None

    def get_siblings(self, skill_uri, max_count=10):
        """Get sibling skills (same broader parent), limited to max_count."""
        siblings = set()
        for parent_uri in self.broader.get(skill_uri, set()):
            for child_uri in self.narrower.get(parent_uri, set()):
                if child_uri != skill_uri:
                    siblings.add(child_uri)
                    if len(siblings) >= max_count:
                        return siblings
        return siblings

    def get_broader_titles(self, skill_uri, depth=1):
        """Get broader (parent category) skill titles."""
        titles = []
        visited = {skill_uri}
        queue = [(skill_uri, 0)]
        while queue:
            current, d = queue.pop(0)
            if d >= depth:
                continue
            for parent_uri in self.broader.get(current, set()):
                if parent_uri not in visited:
                    visited.add(parent_uri)
                    node = self.nodes.get(parent_uri)
                    if node:
                        titles.append(node["title"])
                    queue.append((parent_uri, d + 1))
        return titles

    def get_co_occurring_skills(self, skill_uri, max_count=10):
        """Get skills from occupations that also require this skill."""
        co_skills = set()
        for occ_uri in self.skill_to_occs.get(skill_uri, set()):
            # Essential skills first (more relevant)
            for s_uri in self.occ_essential.get(occ_uri, set()):
                if s_uri != skill_uri:
                    co_skills.add(s_uri)
                    if len(co_skills) >= max_count:
                        return co_skills
            # Then optional (only if we need more)
            if len(co_skills) < max_count:
                for s_uri in self.occ_optional.get(occ_uri, set()):
                    if s_uri != skill_uri:
                        co_skills.add(s_uri)
                        if len(co_skills) >= max_count:
                            return co_skills
        return co_skills

    def get_technology_tools(self, skill_uri, max_count=10):
        """Get O*NET technology tools related to a skill via occupations.

        Path: skill → occupations (ESCO) → O*NET code (crosswalk) → tools
        """
        tools = set()
        for occ_uri in self.skill_to_occs.get(skill_uri, set()):
            for tool_uri in self.occ_tools.get(occ_uri, set()):
                tools.add(tool_uri)
                if len(tools) >= max_count:
                    return tools
        return tools

    def expand_query_terms(self, skill_names, max_terms=20,
                           include_siblings=True,
                           include_broader=True,
                           include_co_occurring=False,
                           include_tools=False):
        """
        Core method: Expand a list of skill names using graph knowledge.

        Args:
            skill_names: list of skill name strings from CV
            max_terms: maximum total expanded terms
            include_siblings: add sibling skills (same parent category)
            include_broader: add broader category names
            include_co_occurring: add co-occurring skills from occupations
            include_tools: add O*NET technology tools from occupations

        Returns:
            list[str] - expanded skill name strings
        """
        if not self._loaded:
            self.load()

        expanded = set()

        for skill_name in skill_names:
            uri = self.find_uri(skill_name)
            if not uri:
                continue

            # 1. Broader categories (always useful)
            if include_broader:
                broader_titles = self.get_broader_titles(uri, depth=1)
                for t in broader_titles:
                    expanded.add(t)
                    if len(expanded) >= max_terms:
                        return list(expanded)

            # 2. Sibling skills (same category)
            if include_siblings:
                siblings = self.get_siblings(uri, max_count=5)
                for s_uri in siblings:
                    node = self.nodes.get(s_uri)
                    if node:
                        expanded.add(node["title"])
                    if len(expanded) >= max_terms:
                        return list(expanded)

            # 3. Co-occurring skills via occupations (optional, heavier)
            if include_co_occurring:
                co_skills = self.get_co_occurring_skills(uri, max_count=5)
                for s_uri in co_skills:
                    node = self.nodes.get(s_uri)
                    if node:
                        expanded.add(node["title"])
                    if len(expanded) >= max_terms:
                        return list(expanded)

            # 4. O*NET technology tools via occupations
            if include_tools:
                tool_uris = self.get_technology_tools(uri, max_count=5)
                for t_uri in tool_uris:
                    node = self.nodes.get(t_uri)
                    if node:
                        expanded.add(node["title"])
                    if len(expanded) >= max_terms:
                        return list(expanded)

        return list(expanded)

    def expand_skills_text(self, skills_text, max_terms=15):
        """
        Convenience method: expand comma-separated skills text.

        Args:
            skills_text: "Python, SQL, React"

        Returns:
            str: "Python, SQL, React, computer programming, JavaScript, ..."
        """
        if not skills_text or not skills_text.strip():
            return skills_text

        original_skills = [s.strip() for s in skills_text.split(",") if s.strip()]
        graph_terms = self.expand_query_terms(
            original_skills,
            max_terms=max_terms,
            include_siblings=True,
            include_broader=True,
            include_co_occurring=False,
        )

        # Merge: original first, then graph expansions (deduped)
        seen = set(s.lower() for s in original_skills)
        result = list(original_skills)
        for term in graph_terms:
            if term.lower() not in seen:
                result.append(term)
                seen.add(term.lower())

        return ", ".join(result)


def get_skill_graph(data_dir=None):
    """Singleton getter — loads graph on first call."""
    global _graph_instance
    if _graph_instance is None:
        _graph_instance = SkillGraph(data_dir=data_dir)
        _graph_instance.load()
    return _graph_instance


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    graph = get_skill_graph()

    # Demo
    test_cases = [
        "Python, SQL, React",
        "project management, agile, scrum",
        "data analysis, machine learning",
        "Java, Spring Boot, microservices",
    ]

    for skills in test_cases:
        expanded = graph.expand_skills_text(skills)
        original_count = len(skills.split(","))
        expanded_count = len(expanded.split(","))
        print(f"\n  Input ({original_count} terms): {skills}")
        print(f"  Output ({expanded_count} terms): {expanded}")
